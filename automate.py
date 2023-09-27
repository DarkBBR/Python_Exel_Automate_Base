#======= PASSO A PASSO DO PROJETO
#=== Passo 1: Importar a biblioteca Openpyxl
#=== Passo 2: definir a variável
#=== Passo 3: Criar uma tabela
#=== Passo 4: Separar e alinhar a tabela
#=== Passo 5: Criar o arquivo e iniciar

import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet("Tabela")

tabela_page = book['Tabela']
tabela_page.append(["Frutas", "Quantidade", "Preço"])
tabela_page.append(["Banana", "5", "R$2,90"])
tabela_page.append(["Maça", "6", "R$4,90"])
tabela_page.append(["Kiwi", "7", "R$8,90"])
tabela_page.append(["Pera", "9", "R$10,90"])
tabela_page.append(["Uva", "20", "R$19,90"])

book.save("Tabela de compras.xlsx")